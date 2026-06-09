"""
XLSX Report Generator - Institutional-grade Excel workbook
Tabs: Summary | Financials | Forensic Scores | Red Flags | Beneish | Altman | Peer | Charts
"""

from __future__ import annotations
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from loguru import logger
from config import REPORT_CONFIG


# Color palette
DARK_BLUE = "1A237E"
RED = "B71C1C"
GREEN = "1B5E20"
ORANGE = "E65100"
LIGHT_BLUE = "E3F2FD"
LIGHT_RED = "FFEBEE"
LIGHT_GREEN = "E8F5E9"
WHITE = "FFFFFF"
DARK_GRAY = "37474F"
LIGHT_GRAY = "ECEFF1"


def header_style(ws, cell_ref, value, bold=True, bg=DARK_BLUE, font_color=WHITE, size=11):
    cell = ws[cell_ref]
    cell.value = value
    cell.font = Font(bold=bold, color=font_color, size=size, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def data_cell(ws, cell_ref, value, bold=False, color=None, number_format=None, align="left"):
    cell = ws[cell_ref]
    cell.value = value
    cell.font = Font(bold=bold, color=color or "000000", name="Calibri", size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if number_format:
        cell.number_format = number_format
    return cell


class XLSXGenerator:
    """Generates multi-tab XLSX forensic investigation workbook."""

    def generate(self, report_data: dict, output_path: Path) -> Path:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        meta = report_data.get("metadata", {})
        company = meta.get("company_name", "Unknown")

        self._create_summary_tab(wb, report_data)
        self._create_financials_tab(wb, report_data)
        self._create_forensic_scores_tab(wb, report_data)
        self._create_red_flags_tab(wb, report_data)
        self._create_beneish_tab(wb, report_data)
        self._create_altman_tab(wb, report_data)
        self._create_risk_dashboard_tab(wb, report_data)
        self._create_management_questions_tab(wb, report_data)
        self._create_monitoring_tab(wb, report_data)

        wb.save(str(output_path))
        logger.info(f"XLSX saved: {output_path.name}")
        return output_path

    def _create_summary_tab(self, wb: openpyxl.Workbook, data: dict) -> None:
        ws = wb.create_sheet("Executive Summary")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 20

        meta = data.get("metadata", {})
        exec_sum = data.get("executive_summary", {})
        company = meta.get("company_name", "Unknown")

        # Title
        ws.merge_cells("A1:C1")
        header_style(ws, "A1", f"FORENSIC INVESTIGATION REPORT - {company.upper()}", size=14)
        ws.row_dimensions[1].height = 40

        ws.merge_cells("A2:C2")
        header_style(ws, "A2", f"{REPORT_CONFIG.firm_name} | {datetime.now().strftime('%B %Y')}", bg=DARK_GRAY, size=10)

        # Risk Score
        score = exec_sum.get("overall_risk_score", 50)
        score_color = RED if score > 60 else (ORANGE if score > 40 else GREEN)
        ws.merge_cells("A4:C4")
        ws.row_dimensions[4].height = 50
        cell = ws["A4"]
        cell.value = f"OVERALL RISK SCORE: {score:.1f}/100"
        cell.font = Font(bold=True, size=18, color=WHITE, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=score_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Verdict
        verdict = exec_sum.get("verdict", "MONITOR")
        ws.merge_cells("A5:C5")
        cell = ws["A5"]
        cell.value = f"INVESTMENT VERDICT: {verdict}"
        cell.font = Font(bold=True, size=14, color=score_color, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Key Metrics
        row = 7
        fields = [
            ("Total Red Flags", exec_sum.get("total_red_flags", 0), RED),
            ("Total Green Flags", exec_sum.get("total_green_flags", 0), GREEN),
            ("Overall Risk Score", f"{score:.1f}/100", score_color),
        ]
        for label, value, color in fields:
            header_style(ws, f"A{row}", label, bg=DARK_GRAY)
            cell = ws[f"B{row}"]
            cell.value = value
            cell.font = Font(bold=True, color=color, size=12, name="Calibri")
            cell.alignment = Alignment(horizontal="center")
            row += 1

        # Risk Components
        row += 1
        header_style(ws, f"A{row}", "RISK COMPONENT SCORES", bg=DARK_BLUE)
        ws.merge_cells(f"A{row}:C{row}")
        row += 1

        agent_analyses = data.get("agent_analyses", {})
        for agent_id, analysis in agent_analyses.items():
            agent_score = analysis.get("risk_score", 50)
            score_color_cell = RED if agent_score > 60 else (ORANGE if agent_score > 40 else GREEN)
            data_cell(ws, f"A{row}", analysis.get("name", ""), bold=True)
            cell = ws[f"B{row}"]
            cell.value = f"{agent_score:.1f}/100"
            cell.font = Font(color=score_color_cell, bold=True)
            data_cell(ws, f"C{row}", analysis.get("red_flag_count", 0), align="center")
            row += 1

        ws.freeze_panes = "A3"

    def _create_financials_tab(self, wb, data: dict) -> None:
        ws = wb.create_sheet("Financial Statements")
        ws.sheet_view.showGridLines = False

        financial_data = data.get("financial_data", {})
        if not financial_data:
            ws["A1"] = "No financial data available"
            return

        years = sorted(financial_data.keys(), reverse=True)

        # Headers
        header_style(ws, "A1", "Metric", bg=DARK_BLUE)
        ws.column_dimensions["A"].width = 35
        for col, year in enumerate(years[:6], 2):
            col_letter = get_column_letter(col)
            header_style(ws, f"{col_letter}1", f"FY{year}", bg=DARK_BLUE)
            ws.column_dimensions[col_letter].width = 18

        # Financial metric sections
        sections = {
            "INCOME STATEMENT": [
                ("Revenue", "revenue"),
                ("Cost of Goods Sold", "cogs"),
                ("Gross Profit", "gross_profit"),
                ("Gross Margin %", None),  # Calculated
                ("EBITDA", "ebitda"),
                ("EBITDA Margin %", None),
                ("EBIT", "ebit"),
                ("Profit Before Tax", "pbt"),
                ("Net Income / PAT", "net_income"),
                ("Net Margin %", None),
                ("EPS (Basic)", "eps"),
                ("EPS (Diluted)", "eps_diluted"),
            ],
            "BALANCE SHEET": [
                ("Total Assets", "total_assets"),
                ("Current Assets", "current_assets"),
                ("Cash & Equivalents", "cash_equivalents"),
                ("Accounts Receivable", "accounts_receivable"),
                ("Inventory", "inventory"),
                ("PPE (Net)", "ppe_net"),
                ("Goodwill", "goodwill"),
                ("Total Liabilities", "total_liabilities"),
                ("Current Liabilities", "current_liabilities"),
                ("Accounts Payable", "accounts_payable"),
                ("Total Debt", "total_debt"),
                ("Shareholders Equity", "shareholder_equity"),
                ("Retained Earnings", "retained_earnings"),
            ],
            "CASH FLOW": [
                ("CFO (Operating Cash)", "cfo"),
                ("CapEx", "capex"),
                ("Free Cash Flow", "fcf"),
                ("Dividends Paid", "dividends_paid"),
            ],
            "KEY RATIOS": [
                ("Current Ratio", "current_ratio"),
                ("Quick Ratio", "quick_ratio"),
                ("D/E Ratio", "de_ratio"),
                ("Interest Coverage", "interest_coverage"),
                ("ROE %", "roe"),
                ("ROA %", "roa"),
                ("ROCE %", "roce"),
                ("Asset Turnover", "asset_turnover"),
                ("Receivable Days (DSO)", "receivable_days"),
                ("Inventory Days (DIO)", "inventory_days"),
                ("Payable Days (DPO)", "payable_days"),
                ("Cash Conversion Cycle", "cash_conversion_cycle"),
            ],
        }

        row = 2
        for section_name, metrics in sections.items():
            # Section header
            ws.merge_cells(f"A{row}:{get_column_letter(len(years[:6])+1)}{row}")
            cell = ws[f"A{row}"]
            cell.value = section_name
            cell.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
            cell.fill = PatternFill("solid", fgColor=DARK_GRAY)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 20
            row += 1

            for label, field_key in metrics:
                cell = ws[f"A{row}"]
                cell.value = label
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="left", vertical="center")

                for col, year in enumerate(years[:6], 2):
                    year_data = financial_data.get(year, {})
                    col_letter = get_column_letter(col)

                    if field_key is None:
                        # Calculate margins
                        if "Gross Margin" in label:
                            rev = year_data.get("revenue", 0)
                            gp = year_data.get("gross_profit", 0)
                            val = (gp / rev * 100) if rev else None
                        elif "EBITDA Margin" in label:
                            rev = year_data.get("revenue", 0)
                            ebitda = year_data.get("ebitda", 0)
                            val = (ebitda / rev * 100) if rev else None
                        elif "Net Margin" in label:
                            rev = year_data.get("revenue", 0)
                            ni = year_data.get("net_income", 0)
                            val = (ni / rev * 100) if rev else None
                        else:
                            val = None
                    else:
                        val = year_data.get(field_key)

                    target_cell = ws[f"{col_letter}{row}"]
                    if val is not None:
                        target_cell.value = val
                        if "%" in label:
                            target_cell.number_format = "0.0%"
                        elif "Days" in label or "Ratio" in label or "Coverage" in label or "Turnover" in label:
                            target_cell.number_format = "0.0"
                        else:
                            target_cell.number_format = "#,##0"
                        target_cell.font = Font(name="Calibri", size=10)
                        target_cell.alignment = Alignment(horizontal="right")
                row += 1

        ws.freeze_panes = "B2"

    def _create_forensic_scores_tab(self, wb, data: dict) -> None:
        ws = wb.create_sheet("Forensic Scores")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 40

        header_style(ws, "A1", "Forensic Model", bg=DARK_BLUE)
        header_style(ws, "B1", "Score", bg=DARK_BLUE)
        header_style(ws, "C1", "Risk Level", bg=DARK_BLUE)
        header_style(ws, "D1", "Interpretation", bg=DARK_BLUE)

        forensic_scores = data.get("forensic_scores", [])
        if forensic_scores:
            score_data = forensic_scores[0]  # Latest year

            rows = [
                ("Beneish M-Score", score_data.get("beneish_m_score"), -1.78,
                 "Manipulation likely if > -1.78"),
                ("Altman Z-Score", score_data.get("altman_z_score"), 1.81,
                 "Distress if < 1.81 | Grey < 2.99 | Safe > 2.99"),
                ("Piotroski F-Score", score_data.get("piotroski_f_score"), 3,
                 "Strong: 7-9 | Neutral: 3-6 | Weak: 0-2"),
                ("Dechow F-Score", score_data.get("dechow_f_score"), 0.01,
                 "Misreporting probability. High risk > 0.025"),
                ("Accrual Ratio (CF)", score_data.get("accrual_ratio"), 0.05,
                 "Earnings quality concern if > 0.10"),
                ("Cash Earnings Ratio", score_data.get("cash_earnings_ratio"), 0.8,
                 "CFO/NI. Should be > 0.85 for quality earnings"),
                ("Overall Risk Score", score_data.get("overall_risk_score"), 50,
                 "0=Very Low Risk | 100=Extreme Risk"),
            ]

            for i, (name, score, threshold, interpretation) in enumerate(rows, 2):
                data_cell(ws, f"A{i}", name, bold=True)
                if score is not None:
                    cell = ws[f"B{i}"]
                    cell.value = round(score, 4)
                    cell.number_format = "0.0000"
                    cell.alignment = Alignment(horizontal="center")

                    # Color based on risk
                    is_risky = False
                    if name == "Beneish M-Score":
                        is_risky = score > threshold
                    elif name == "Altman Z-Score":
                        is_risky = score < threshold
                    elif name == "Piotroski F-Score":
                        is_risky = score < threshold
                    else:
                        is_risky = score > threshold

                    risk_text = "HIGH RISK" if is_risky else "ACCEPTABLE"
                    cell_c = ws[f"C{i}"]
                    cell_c.value = risk_text
                    cell_c.font = Font(bold=True, color=RED if is_risky else GREEN, name="Calibri")
                    cell_c.alignment = Alignment(horizontal="center")

                data_cell(ws, f"D{i}", interpretation)

    def _create_red_flags_tab(self, wb, data: dict) -> None:
        ws = wb.create_sheet("Red Flags Registry")
        ws.sheet_view.showGridLines = False

        headers = ["#", "Severity", "Agent", "Finding Title", "Evidence", "Fiscal Year"]
        widths = [5, 12, 25, 45, 60, 10]
        for col, (hdr, width) in enumerate(zip(headers, widths), 1):
            col_letter = get_column_letter(col)
            header_style(ws, f"{col_letter}1", hdr, bg=RED)
            ws.column_dimensions[col_letter].width = width

        red_flags = data.get("red_flags", [])
        severity_order = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
        sorted_flags = sorted(red_flags, key=lambda x: severity_order.get(x.get("risk_level", "LOW"), 3))

        for i, flag in enumerate(sorted_flags, 2):
            severity = flag.get("risk_level", "HIGH")
            bg = "FF0000" if severity == "CRITICAL" else ("FF6600" if severity == "HIGH" else "FFCC00")

            ws[f"A{i}"].value = i - 1
            ws[f"B{i}"].value = severity
            ws[f"B{i}"].font = Font(bold=True, color=bg, name="Calibri")
            ws[f"C{i}"].value = flag.get("agent", "")
            ws[f"D{i}"].value = flag.get("title", "")
            ws[f"D{i}"].font = Font(bold=True, name="Calibri", size=10)
            ws[f"E{i}"].value = flag.get("evidence", "")[:200]
            ws[f"E{i}"].alignment = Alignment(wrap_text=True)
            ws[f"F{i}"].value = flag.get("fiscal_year", "")

            # Alternate row colors
            if i % 2 == 0:
                for col in range(1, 7):
                    ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=LIGHT_RED)

    def _create_beneish_tab(self, wb, data: dict) -> None:
        ws = wb.create_sheet("Beneish M-Score")
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 50

        # Title
        ws.merge_cells("A1:D1")
        header_style(ws, "A1", "BENEISH M-SCORE - EARNINGS MANIPULATION MODEL", size=12)

        # Description
        ws["A2"] = "Model: Beneish (1999) | Threshold: M-Score > -1.78 → Likely Manipulator"
        ws["A2"].font = Font(italic=True, name="Calibri", size=10)

        header_style(ws, "A3", "Component", bg=DARK_GRAY)
        header_style(ws, "B3", "Value", bg=DARK_GRAY)
        header_style(ws, "C3", "Concern Level", bg=DARK_GRAY)
        header_style(ws, "D3", "What It Measures", bg=DARK_GRAY)

        forensic_scores = data.get("forensic_scores", [{}])
        scores = forensic_scores[0] if forensic_scores else {}

        components = [
            ("DSRI - Days Sales Receivable Index", scores.get("beneish_dsri"), 1.031,
             "Receivables growing faster than sales → revenue inflation risk"),
            ("GMI - Gross Margin Index", scores.get("beneish_gmi"), 1.014,
             "Gross margin deterioration → earnings pressure"),
            ("AQI - Asset Quality Index", scores.get("beneish_aqi"), 1.039,
             "More assets in soft assets → manipulation latitude"),
            ("SGI - Sales Growth Index", scores.get("beneish_sgi"), 1.600,
             "High growth companies more likely to manipulate"),
            ("DEPI - Depreciation Index", scores.get("beneish_depi"), 1.001,
             "Declining depreciation rate → extending asset lives"),
            ("SGAI - SG&A Index", scores.get("beneish_sgai"), 1.054,
             "Disproportionate overhead growth"),
            ("LVGI - Leverage Index", scores.get("beneish_lvgi"), 1.111,
             "Rising leverage → debt covenant pressure"),
            ("TATA - Total Accruals / Assets", scores.get("beneish_tata"), 0.031,
             "Accruals-based earnings not backed by cash"),
            ("", None, None, ""),
            ("M-SCORE (TOTAL)", scores.get("beneish_m_score"), -1.78,
             "< -1.78 = Clean | > -1.78 = Manipulator"),
        ]

        for i, (comp, val, threshold, desc) in enumerate(components, 4):
            ws[f"A{i}"].value = comp
            ws[f"A{i}"].font = Font(bold=(comp == "M-SCORE (TOTAL)"), name="Calibri", size=10)

            if val is not None and threshold is not None:
                ws[f"B{i}"].value = round(val, 4)
                ws[f"B{i}"].number_format = "0.0000"
                ws[f"B{i}"].alignment = Alignment(horizontal="center")

                is_risky = (val > threshold) if comp != "M-SCORE (TOTAL)" else (val > threshold)
                concern_text = "⚠️ CONCERN" if is_risky else "✅ Normal"
                ws[f"C{i}"].value = concern_text
                ws[f"C{i}"].font = Font(color=RED if is_risky else GREEN, bold=True, name="Calibri")
                ws[f"C{i}"].alignment = Alignment(horizontal="center")

                if comp == "M-SCORE (TOTAL)":
                    fill_color = RED if val > -1.78 else GREEN
                    for col in ["A", "B", "C", "D"]:
                        ws[f"{col}{i}"].fill = PatternFill("solid", fgColor="FFEBEE" if val > -1.78 else "E8F5E9")

            ws[f"D{i}"].value = desc
            ws[f"D{i}"].alignment = Alignment(wrap_text=True)

    def _create_altman_tab(self, wb, data: dict) -> None:
        ws = wb.create_sheet("Altman Z-Score")
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 50

        ws.merge_cells("A1:C1")
        header_style(ws, "A1", "ALTMAN Z-SCORE - BANKRUPTCY PREDICTION MODEL", size=12)

        header_style(ws, "A2", "Variable", bg=DARK_GRAY)
        header_style(ws, "B2", "Value", bg=DARK_GRAY)
        header_style(ws, "C2", "Description", bg=DARK_GRAY)

        forensic_scores = data.get("forensic_scores", [{}])
        scores = forensic_scores[0] if forensic_scores else {}
        z_score = scores.get("altman_z_score", 0)
        zone = scores.get("altman_zone", "UNKNOWN")

        components = [
            ("X1 = WC / Total Assets", scores.get("altman_x1"), "Working capital / total assets"),
            ("X2 = RE / Total Assets", scores.get("altman_x2"), "Retained earnings / total assets"),
            ("X3 = EBIT / Total Assets", scores.get("altman_x3"), "EBIT / total assets"),
            ("X4 = Equity / Total Liabilities", scores.get("altman_x4"), "Market cap / total liabilities"),
            ("X5 = Revenue / Total Assets", scores.get("altman_x5"), "Asset turnover"),
        ]

        for i, (name, val, desc) in enumerate(components, 3):
            ws[f"A{i}"].value = name
            if val is not None:
                ws[f"B{i}"].value = round(val, 4)
                ws[f"B{i}"].number_format = "0.0000"
                ws[f"B{i}"].alignment = Alignment(horizontal="center")
            ws[f"C{i}"].value = desc

        # Z-Score row
        row = len(components) + 4
        ws.merge_cells(f"A{row}:C{row}")
        zone_color = RED if zone == "DISTRESS" else (ORANGE if zone == "GREY" else GREEN)
        cell = ws[f"A{row}"]
        cell.value = f"Z-SCORE: {z_score:.3f} → {zone} ZONE"
        cell.font = Font(bold=True, color=WHITE, size=12, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=zone_color)
        cell.alignment = Alignment(horizontal="center")

        # Zone reference
        row += 2
        ws[f"A{row}"].value = "Zone Reference"
        ws[f"A{row}"].font = Font(bold=True, name="Calibri")
        for zone_name, lower, upper, color in [
            ("SAFE ZONE", "2.99", ">", GREEN),
            ("GREY ZONE", "1.81", "1.81 – 2.99", ORANGE),
            ("DISTRESS ZONE", "<1.81", "High bankruptcy risk", RED),
        ]:
            row += 1
            ws[f"A{row}"].value = zone_name
            ws[f"A{row}"].font = Font(color=color, bold=True, name="Calibri")
            ws[f"B{row}"].value = lower
            ws[f"C{row}"].value = upper

    def _create_risk_dashboard_tab(self, wb, data: dict) -> None:
        ws = wb.create_sheet("Risk Dashboard")
        ws.merge_cells("A1:E1")
        header_style(ws, "A1", "RISK SCORE DASHBOARD", size=14)

        exec_sum = data.get("executive_summary", {})
        overall_score = exec_sum.get("overall_risk_score", 50)

        ws.merge_cells("A3:E3")
        score_color = RED if overall_score > 60 else (ORANGE if overall_score > 40 else GREEN)
        cell = ws["A3"]
        cell.value = f"OVERALL RISK SCORE: {overall_score:.1f} / 100"
        cell.font = Font(bold=True, color=WHITE, size=16, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=score_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 40

    def _create_management_questions_tab(self, wb, data: dict) -> None:
        ws = wb.create_sheet("Management Questions")
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 80

        ws.merge_cells("A1:B1")
        header_style(ws, "A1", "MANAGEMENT QUESTIONNAIRE (50-80 Questions)", size=12)

        questions = data.get("management_questions", [])
        if not questions:
            questions = ["No management questions generated. Run full investigation for complete questionnaire."]

        for i, q in enumerate(questions, 2):
            ws[f"A{i}"].value = i - 1
            ws[f"A{i}"].font = Font(bold=True, name="Calibri", size=10)
            ws[f"B{i}"].value = q
            ws[f"B{i}"].alignment = Alignment(wrap_text=True)
            ws[f"B{i}"].font = Font(name="Calibri", size=10)
            if i % 2 == 0:
                ws[f"B{i}"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    def _create_monitoring_tab(self, wb, data: dict) -> None:
        ws = wb.create_sheet("Monitoring Framework")
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 35

        ws.merge_cells("A1:C1")
        header_style(ws, "A1", "QUARTERLY MONITORING FRAMEWORK", size=12)

        headers = ["Metric to Monitor", "Trigger Threshold", "Recommended Action"]
        for col, hdr in enumerate(headers, 1):
            header_style(ws, f"{get_column_letter(col)}2", hdr, bg=DARK_GRAY)

        triggers = data.get("monitoring_framework", [])
        for i, trigger in enumerate(triggers, 3):
            ws[f"A{i}"].value = trigger.get("metric", "")
            ws[f"A{i}"].font = Font(bold=True, name="Calibri", size=10)
            ws[f"B{i}"].value = trigger.get("threshold", "")
            ws[f"B{i}"].font = Font(color=RED, name="Calibri", size=10)
            ws[f"C{i}"].value = trigger.get("action", "")
            ws[f"C{i}"].font = Font(name="Calibri", size=10)
